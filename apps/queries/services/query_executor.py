"""
Servicio para ejecutar consultas SQL de forma segura
"""
from django.conf import settings
from django.utils.translation import gettext_lazy as _
from decimal import Decimal
import time
import logging

# NO importar QueryExecution aquí arriba, hacerlo dentro de las funciones
from .query_builder import QueryBuilderService
from .query_validator import QueryValidatorService

logger = logging.getLogger(__name__)


class QueryExecutorService:
    """
    Ejecuta consultas SQL dinámicas de forma segura y controlada.
    
    Maneja timeout, logging y paginación de resultados.
    """
    
    def __init__(self):
        # Import local para evitar circular import
        from ..repositories.sql_server_repository import SQLServerRepository
        
        self.repository = SQLServerRepository()
        self.timeout = getattr(settings, 'SQL_QUERY_TIMEOUT', 30)
        self.max_results = getattr(settings, 'MAX_RESULTS_LIMIT', 10000)
    
    def execute_query(self, query_obj, params_dict=None, page=1, page_size=None, usuario='system'):
        """
        Ejecuta una consulta dinámica con parámetros opcionales.
        
        Args:
            query_obj: Instancia de DynamicQuery
            params_dict: Dict con parámetros {posicion_where: valor}
            page: Número de página (para paginación)
            page_size: Tamaño de página (None usa RESULTS_PER_PAGE del settings)
            usuario: Nombre del usuario ejecutando la consulta
        
        Returns:
            dict: Resultado completo de la ejecución
        """
        start_time = time.time()
        
        result = {
            'success': False,
            'data': [],
            'columns': [],
            'total_rows': 0,
            'execution_time': 0.0,
            'error': None,
            'sql': '',
            'page': page,
            'page_size': page_size or getattr(settings, 'RESULTS_PER_PAGE', 50),
            'has_next': False,
            'has_previous': page > 1
        }
        
        try:
            # Validar parámetros requeridos
            validation = QueryBuilderService.validate_parameters(query_obj, params_dict or {})
            if not validation['valid']:
                result['error'] = '; '.join(validation['errors'])
                return result
            
            # Construir SQL
            sql = QueryBuilderService.build_query(query_obj, params_dict)
            result['sql'] = sql
            
            # Ejecutar con paginación
            execution_result = self.repository.execute_query_with_pagination(
                sql=sql,
                page=page,
                page_size=result['page_size'],
                timeout=self.timeout
            )
            
            if execution_result['success']:
                result['success'] = True
                result['data'] = execution_result['data']
                result['columns'] = execution_result['columns']
                result['total_rows'] = execution_result['total_rows']
                result['has_next'] = execution_result['has_next']
                
                # Si no hay columnas en el resultado, usar aliases del SELECT
                if not result['columns'] and result['data']:
                    result['columns'] = self._extract_columns_from_select(query_obj.select_clause)
            else:
                result['error'] = execution_result['error']
            
        except Exception as e:
            logger.error(f"Error ejecutando consulta {query_obj.nombre}: {str(e)}")
            result['error'] = str(e)
        
        finally:
            # Calcular tiempo de ejecución
            result['execution_time'] = round(time.time() - start_time, 3)
            
            # Registrar ejecución
            self._log_execution(
                query_obj=query_obj,
                usuario=usuario,
                params_dict=params_dict or {},
                result=result
            )
        
        return result
    
    def execute_test_query(self, query_obj, test_params=None, usuario='system'):
        """
        Ejecuta una consulta en modo de prueba (máximo 10 resultados).
        """
        if test_params is None:
            test_params = self._generate_test_params(query_obj)
        
        return self.execute_query(
            query_obj=query_obj,
            params_dict=test_params,
            page=1,
            page_size=10,
            usuario=usuario
        )
    
    def _generate_test_params(self, query_obj):
        """Genera parámetros de prueba automáticos."""
        test_params = {}
        
        for param in query_obj.parametros.all():
            if param.valor_por_defecto:
                test_params[param.posicion_where] = param.valor_por_defecto
            else:
                test_params[param.posicion_where] = self._get_test_value_by_type(param.tipo_dato)
        
        return test_params
    
    def _get_test_value_by_type(self, tipo_dato):
        """Retorna un valor de prueba según el tipo de dato."""
        test_values = {
            'texto': 'TEST',
            'numero': 1,
            'decimal': 1.0,
            'fecha': '2025-01-01',
            'boolean': True
        }
        return test_values.get(tipo_dato, 'TEST')
    
    def _extract_columns_from_select(self, select_clause):
        """Extrae nombres de columnas desde la cláusula SELECT."""
        import re
        
        columns = []
        pattern = r'\bAS\s+([a-zA-Z_][\w]*)'
        matches = re.findall(pattern, select_clause, re.IGNORECASE)
        
        if matches:
            columns = matches
        else:
            parts = select_clause.split(',')
            for part in parts:
                field = part.strip().split('.')[-1].split()[-1]
                columns.append(field)
        
        return columns
    
    def _log_execution(self, query_obj, usuario, params_dict, result):
        """Registra la ejecución en la base de datos."""
        # Import local para evitar circular import
        from ..models import QueryExecution
        
        try:
            QueryExecution.crear_log(
                query=query_obj,
                usuario=usuario,
                parametros=params_dict,
                total_resultados=result['total_rows'],
                tiempo_ejecucion=Decimal(str(result['execution_time'])),
                exitosa=result['success'],
                mensaje_error=result['error'] or '',
                sql_ejecutado=result['sql']
            )
        except Exception as e:
            logger.error(f"Error registrando ejecución: {str(e)}")
    
    def export_to_excel(self, query_obj, params_dict=None, usuario='system'):
        """Exporta los resultados a Excel."""
        from openpyxl import Workbook
        from django.conf import settings
        import os
        from datetime import datetime
        
        result = {
            'success': False,
            'file_path': None,
            'error': None
        }
        
        try:
            execution_result = self.execute_query(
                query_obj=query_obj,
                params_dict=params_dict,
                page=1,
                page_size=self.max_results,
                usuario=usuario
            )
            
            if not execution_result['success']:
                result['error'] = execution_result['error']
                return result
            
            wb = Workbook()
            ws = wb.active
            ws.title = query_obj.nombre[:31]
            
            headers = execution_result['columns']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            for row_num, row_data in enumerate(execution_result['data'], 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            media_root = settings.MEDIA_ROOT
            exports_dir = os.path.join(media_root, 'exports')
            os.makedirs(exports_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{query_obj.nombre}_{timestamp}.xlsx"
            file_path = os.path.join(exports_dir, filename)
            
            wb.save(file_path)
            
            result['success'] = True
            result['file_path'] = file_path
            result['filename'] = filename
            
        except Exception as e:
            logger.error(f"Error exportando a Excel: {str(e)}")
            result['error'] = str(e)
        
        return result
    
    def export_to_csv(self, query_obj, params_dict=None, usuario='system'):
        """Exporta los resultados a CSV."""
        import csv
        from django.conf import settings
        import os
        from datetime import datetime
        
        result = {
            'success': False,
            'file_path': None,
            'error': None
        }
        
        try:
            execution_result = self.execute_query(
                query_obj=query_obj,
                params_dict=params_dict,
                page=1,
                page_size=self.max_results,
                usuario=usuario
            )
            
            if not execution_result['success']:
                result['error'] = execution_result['error']
                return result
            
            media_root = settings.MEDIA_ROOT
            exports_dir = os.path.join(media_root, 'exports')
            os.makedirs(exports_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{query_obj.nombre}_{timestamp}.csv"
            file_path = os.path.join(exports_dir, filename)
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(execution_result['columns'])
                
                for row in execution_result['data']:
                    writer.writerow(row)
            
            result['success'] = True
            result['file_path'] = file_path
            result['filename'] = filename
            
        except Exception as e:
            logger.error(f"Error exportando a CSV: {str(e)}")
            result['error'] = str(e)
        
        return result